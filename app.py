# app.py
from flask import Flask, render_template, request, redirect, url_for, jsonify
from openpyxl import load_workbook
from datetime import datetime
import os

app = Flask(__name__)

EXCEL_PATH = "plantilla.xlsx"

# Leer precios fijos desde hoja de referencia
def obtener_precio_paquete(paquete):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    if "precios" not in wb.sheetnames:
        return 0
    ws = wb["precios"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[0].strip().lower() == paquete.strip().lower():
            return float(row[1])
    return 0

@app.route('/')
def menu():
    return render_template('menu.html')

@app.route('/ventas')
def vista_ventas():
    return render_template('ventas.html')

@app.route('/gastos')
def vista_gastos():
    return render_template('gastos.html')

@app.route('/reporte')
def reporte():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws_ventas = wb["ventas"]
    ws_gastos = wb["gastos"]

    total_ventas = 0
    total_pago = 0
    total_pendiente = 0
    deudores = []

    for row in ws_ventas.iter_rows(min_row=2, values_only=True):
        total_ventas += float(row[4] or 0)
        total_pago += float(row[5] or 0)
        pendiente = float(row[7] or 0)
        total_pendiente += pendiente
        if pendiente > 0:
            deudores.append({"cliente": row[2], "pendiente": pendiente})

    total_gastos_fijos = 0
    total_gastos_variables = 0

    for row in ws_gastos.iter_rows(min_row=2, values_only=True):
        tipo = row[2]
        valor_total = float(row[6] or 0)
        if tipo == "fijo":
            total_gastos_fijos += valor_total
        elif tipo == "variable":
            total_gastos_variables += valor_total

    utilidad = total_pago - total_gastos_variables

    return render_template("reporte.html",
        total_ventas=total_ventas,
        total_pago=total_pago,
        total_pendiente=total_pendiente,
        deudores=deudores,
        gastos_fijos=total_gastos_fijos,
        gastos_variables=total_gastos_variables,
        utilidad=utilidad
    )

@app.route('/registrar_venta', methods=['POST'])
def registrar_venta():
    cliente = request.form['cliente']
    paquete = request.form['paquete']
    valor_producto = obtener_precio_paquete(paquete)
    pago = float(request.form['pago'])
    medio_pago = request.form['medio_pago']
    pendiente = valor_producto - pago
    observaciones = request.form.get('observaciones', '')

    wb = load_workbook(EXCEL_PATH)
    ws = wb["ventas"]

    next_id = ws.max_row if ws.max_row > 1 else 1
    nueva_fila = [
        next_id,
        datetime.today().strftime('%Y-%m-%d'),
        cliente,
        paquete,
        valor_producto,
        pago,
        medio_pago,
        pendiente,
        observaciones
    ]
    ws.append(nueva_fila)
    wb.save(EXCEL_PATH)

    return redirect(url_for('vista_ventas'))

@app.route('/registrar_gasto', methods=['POST'])
def registrar_gasto():
    descripcion = request.form['descripcion']
    tipo_gasto = request.form['tipo_gasto']
    cantidad = int(request.form['cantidad'])
    valor_unitario = float(request.form['valor_unitario'])
    valor_total = cantidad * valor_unitario

    wb = load_workbook(EXCEL_PATH)
    ws = wb["gastos"]

    next_id = ws.max_row if ws.max_row > 1 else 1
    nueva_fila = [
        next_id,
        datetime.today().strftime('%Y-%m-%d'),
        tipo_gasto,
        descripcion,
        cantidad,
        valor_unitario,
        valor_total
    ]
    ws.append(nueva_fila)
    wb.save(EXCEL_PATH)

    return redirect(url_for('vista_gastos'))

@app.route('/precio/<paquete>')
def obtener_precio(paquete):
    precio = obtener_precio_paquete(paquete)
    return jsonify({"precio": precio})

if __name__ == '__main__':
    app.run(debug=True)